import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

OFFICIAL_FILE = "Clean_Ramat_Gan_Pilot.xlsx"
OUTPUT_FILE = "Parent_Input_Data_Structure.xlsx"

official = pd.read_excel(OFFICIAL_FILE, sheet_name="Cleaned_Data")

# --------------------------------------------------
# 1. Garden Profile
# טבלת הבסיס של כל הגנים
# --------------------------------------------------
garden_profile = pd.DataFrame({
    "garden_id": official["id"],
    "name": official["name"],
    "city": official["city"],
    "address": official["address"],
    "official_phone": official["phone"],
    "ownership": official["ownership"],
    "sector": official["sector"],
    "manager": official["manager"],
    "license_status": official["license_status"],
    "garden_type": "Unknown",
    "ages": "Unknown",
    "activity_hours": "Unknown",
    "friday": "Unknown",
    "external_link": "",

    # מאפייני גן שמתעדכנים רק אחרי בדיקה / מקור אמין
    "education_type": "Unknown",
    "education_type_other": "",
    "religious_orientation": "Unknown",
    "religious_orientation_other": "",
    "nutrition_type": "Unknown",
    "nutrition_type_other": "",
    "has_cameras": "Unknown",
    "cameras_open_to_parents": "Unknown",
    "has_protected_space": "Unknown",

    # שדות ניהול מידע
    "profile_data_status": "official_only",
    "last_verified_at": "",
    "verified_by": ""
})

# --------------------------------------------------
# 2. Parent Reviews Template
# כל שורה מייצגת ביקורת / דיווח של הורה על גן קיים
# --------------------------------------------------
parent_reviews_template = pd.DataFrame(columns=[
    "review_id",
    "garden_id",
    "garden_name",

    "rating_1_to_5",
    "price_monthly",
    "price_includes_friday",
    "price_details",

    "staff_count",
    "children_count",
    "staff_child_ratio",

    "parent_feeling",
    "review_text",

    "created_at",
    "status"
])

# --------------------------------------------------
# 3. Suggested Updates Template
# הצעות לעדכון מאפייני גן
# לא מעדכנות ישירות את פרופיל הגן
# --------------------------------------------------
suggested_updates_template = pd.DataFrame(columns=[
    "suggestion_id",
    "garden_id",
    "garden_name",

    "field_name",
    "suggested_value",
    "source",
    "comment",

    "created_at",
    "status"
])

# --------------------------------------------------
# 4. Lookup Values
# ערכים מותרים לשדות מובנים
# --------------------------------------------------
lookup_lists = {
    "yes_no_unknown": [
        "כן",
        "לא",
        "לא ידוע"
    ],

    "review_status": [
        "pending_review",
        "approved",
        "rejected"
    ],

    "profile_data_status": [
        "official_only",
        "community_reported",
        "externally_verified",
        "admin_verified",
        "official_and_community_reported"
    ],

    "parent_feeling": [
        "מאוד מרוצה",
        "מרוצה",
        "ניטרלי",
        "לא מרוצה",
        "מאוד לא מרוצה"
    ],

    "education_type": [
        "רגיל / כללי",
        "מונטסורי",
        "אנתרופוסופי / וולדורף",
        "דמוקרטי",
        "טבע / יער",
        "רג׳יו אמיליה",
        "דו-לשוני",
        "חינוך מיוחד",
        "שילוב",
        "אומנויות",
        "מוזיקה",
        "מדעים / STEM",
        "ספורט / תנועה",
        "אחר",
        "Unknown"
    ],

    "religious_orientation": [
        "חילוני",
        "מסורתי",
        "דתי",
        "ממלכתי דתי",
        "חרדי",
        "מעורב",
        "אחר",
        "לא ידוע",
        "Unknown"
    ],

    "nutrition_type": [
        "רגילה",
        "כשרה",
        "כשרה למהדרין",
        "צמחונית",
        "טבעונית",
        "ללא גלוטן",
        "ללא אלרגנים",
        "ביתית",
        "קייטרינג",
        "הורה מביא אוכל",
        "אחר",
        "Unknown"
    ],

"suggested_field_name": [
    "display_name",
    "garden_type",
    "ages",
    "activity_hours",
    "friday",
    "external_link",
    "education_type",
    "religious_orientation",
    "nutrition_type",
    "has_cameras",
    "cameras_open_to_parents",
    "has_protected_space"
],

    "source": [
        "parent",
        "external_site",
        "admin",
        "other"
    ]
}

lookup_values = pd.DataFrame({
    key: pd.Series(value)
    for key, value in lookup_lists.items()
})

# --------------------------------------------------
# כתיבה ראשונית לאקסל
# --------------------------------------------------
with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
    garden_profile.to_excel(writer, sheet_name="Garden_Profile", index=False)
    parent_reviews_template.to_excel(writer, sheet_name="Parent_Reviews_Template", index=False)
    suggested_updates_template.to_excel(writer, sheet_name="Suggested_Updates_Template", index=False)
    lookup_values.to_excel(writer, sheet_name="Lookup_Values", index=False)

# --------------------------------------------------
# עיצוב + רשימות בחירה
# --------------------------------------------------
wb = load_workbook(OUTPUT_FILE)

header_fill = PatternFill(start_color="D9EAF7", end_color="D9EAF7", fill_type="solid")

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    ws.freeze_panes = "A2"

    for column_cells in ws.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter

        for cell in column_cells:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[column_letter].width = min(max_length + 2, 35)

# --------------------------------------------------
# פונקציות עזר לרשימות בחירה
# --------------------------------------------------
def get_lookup_range(sheet, column_name):
    header_row = 1

    for cell in sheet[header_row]:
        if cell.value == column_name:
            col_letter = cell.column_letter
            max_row = sheet.max_row

            # למצוא את השורה האחרונה שיש בה ערך בעמודה הזאת
            last_value_row = header_row
            for row in range(header_row + 1, max_row + 1):
                if sheet[f"{col_letter}{row}"].value not in [None, ""]:
                    last_value_row = row

            return f"=Lookup_Values!${col_letter}$2:${col_letter}${last_value_row}"

    raise ValueError(f"Column {column_name} not found in Lookup_Values")


def get_column_letter_by_header(ws, header_name):
    for cell in ws[1]:
        if cell.value == header_name:
            return cell.column_letter

    raise ValueError(f"Column {header_name} not found in sheet {ws.title}")


def add_list_validation(ws, column_name, formula_range, start_row=2, end_row=1000):
    col_letter = get_column_letter_by_header(ws, column_name)

    validation = DataValidation(
        type="list",
        formula1=formula_range,
        allow_blank=True
    )

    ws.add_data_validation(validation)
    validation.add(f"{col_letter}{start_row}:{col_letter}{end_row}")


def add_number_validation(ws, column_name, min_value, max_value, start_row=2, end_row=1000):
    col_letter = get_column_letter_by_header(ws, column_name)

    validation = DataValidation(
        type="whole",
        operator="between",
        formula1=str(min_value),
        formula2=str(max_value),
        allow_blank=True
    )

    ws.add_data_validation(validation)
    validation.add(f"{col_letter}{start_row}:{col_letter}{end_row}")


# --------------------------------------------------
# החלת רשימות בחירה על הגיליונות
# --------------------------------------------------
lookup_ws = wb["Lookup_Values"]
garden_ws = wb["Garden_Profile"]
reviews_ws = wb["Parent_Reviews_Template"]
updates_ws = wb["Suggested_Updates_Template"]

yes_no_range = get_lookup_range(lookup_ws, "yes_no_unknown")
review_status_range = get_lookup_range(lookup_ws, "review_status")
profile_status_range = get_lookup_range(lookup_ws, "profile_data_status")
parent_feeling_range = get_lookup_range(lookup_ws, "parent_feeling")
education_type_range = get_lookup_range(lookup_ws, "education_type")
religious_orientation_range = get_lookup_range(lookup_ws, "religious_orientation")
nutrition_type_range = get_lookup_range(lookup_ws, "nutrition_type")
suggested_field_range = get_lookup_range(lookup_ws, "suggested_field_name")
source_range = get_lookup_range(lookup_ws, "source")

# Garden_Profile validations
add_list_validation(garden_ws, "education_type", education_type_range, end_row=500)
add_list_validation(garden_ws, "religious_orientation", religious_orientation_range, end_row=500)
add_list_validation(garden_ws, "nutrition_type", nutrition_type_range, end_row=500)
add_list_validation(garden_ws, "has_cameras", yes_no_range, end_row=500)
add_list_validation(garden_ws, "cameras_open_to_parents", yes_no_range, end_row=500)
add_list_validation(garden_ws, "has_protected_space", yes_no_range, end_row=500)
add_list_validation(garden_ws, "profile_data_status", profile_status_range, end_row=500)

# Parent_Reviews_Template validations
add_number_validation(reviews_ws, "rating_1_to_5", 1, 5)
add_list_validation(reviews_ws, "price_includes_friday", yes_no_range)
add_list_validation(reviews_ws, "parent_feeling", parent_feeling_range)
add_list_validation(reviews_ws, "status", review_status_range)

# Suggested_Updates_Template validations
add_list_validation(updates_ws, "field_name", suggested_field_range)
add_list_validation(updates_ws, "source", source_range)
add_list_validation(updates_ws, "status", review_status_range)

# --------------------------------------------------
# הערות הסבר בתוך הקובץ
# --------------------------------------------------
notes = [
    {
        "sheet": "Garden_Profile",
        "cell": "A1",
        "note": "טבלת הבסיס של כל הגנים. מאפייני גן כמו סוג חינוך ותזונה לא מתעדכנים ישירות על ידי הורה אלא אחרי בדיקה."
    },
    {
        "sheet": "Parent_Reviews_Template",
        "cell": "A1",
        "note": "כל שורה מייצגת ביקורת או דיווח של הורה על גן קיים."
    },
    {
        "sheet": "Suggested_Updates_Template",
        "cell": "A1",
        "note": "כאן נשמרות הצעות הורים לעדכון מאפייני גן. הן לא מעדכנות את הפרופיל עד אישור."
    },
    {
        "sheet": "Lookup_Values",
        "cell": "A1",
        "note": "ערכים מותרים לשדות מובנים. באתר זה יהפוך ל-dropdown או checkbox."
    }
]

for item in notes:
    ws = wb[item["sheet"]]
    ws[item["cell"]].comment = None

# שמירה
wb.save(OUTPUT_FILE)

print(f"✅ Done | Created {OUTPUT_FILE} with {len(garden_profile)} gardens")